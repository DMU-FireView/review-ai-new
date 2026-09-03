"""Data preparation and deterministic splitting for the P_text baseline."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
import random
import re
from typing import Any, Iterable, Mapping, Sequence

LABEL_MAPPING = {"NORMAL": 0, "SUSPICIOUS": 1}
EXCLUDED_LABELS = frozenset({"UNCERTAIN", "DISCLOSED_PROMO", "INVALID"})


class InsufficientDataError(ValueError):
    """Raised when three meaningful classification splits cannot be made."""


@dataclass(frozen=True)
class PreparedData:
    examples: list[dict[str, Any]]
    raw_selected_count: int
    duplicate_rows_removed: int
    conflicting_groups_removed: int


def normalize_text(text: str) -> str:
    """Normalize whitespace/case solely for duplicate detection."""
    return re.sub(r"\s+", " ", text).strip().casefold()


def prepare_records(
    records: Iterable[Mapping[str, Any]],
    *,
    text_column: str = "review",
    label_column: str = "final_label",
) -> PreparedData:
    """Filter labels and remove duplicate text groups before any split.

    If the same normalized text has contradictory labels, the whole group is
    excluded. Returned model examples deliberately contain only ``text`` and
    ``label``; human-authored metadata never reaches tokenization.
    """
    selected: list[dict[str, Any]] = []
    for row in records:
        name = str(row.get(label_column, "")).strip().upper()
        text = str(row.get(text_column, "")).strip()
        if name in LABEL_MAPPING and text:
            selected.append({"text": text, "label": LABEL_MAPPING[name]})

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in selected:
        groups[normalize_text(row["text"])].append(row)

    examples: list[dict[str, Any]] = []
    conflicts = 0
    removed = 0
    for group in groups.values():
        if len({row["label"] for row in group}) > 1:
            conflicts += 1
            removed += len(group)
            continue
        examples.append(group[0])
        removed += len(group) - 1
    return PreparedData(examples, len(selected), removed, conflicts)


def _allocate_class(items: list[dict[str, Any]]) -> tuple[list, list, list]:
    """Allocate roughly 80/10/10 while guaranteeing each split one item."""
    count = len(items)
    if count < 3:
        raise InsufficientDataError(
            "Each class needs at least 3 unique examples for train/validation/test; "
            f"received {count}. Collect more data or use cross-validation."
        )
    validation = max(1, round(count * 0.1))
    test = max(1, round(count * 0.1))
    train = count - validation - test
    if train < 1:
        raise InsufficientDataError("Not enough examples to keep one item in every split")
    return items[:train], items[train : train + validation], items[train + validation :]


def stratified_split(
    examples: Sequence[Mapping[str, Any]], *, seed: int = 42
) -> dict[str, list[dict[str, Any]]]:
    """Deterministic per-class split, safe for the four-positive dataset.

    This is a fallback to ordinary fractional stratification: rounding a class
    of four can silently leave validation/test without positives. It instead
    guarantees every split contains both labels, or raises a clear error.
    """
    by_label: dict[int, list[dict[str, Any]]] = {0: [], 1: []}
    for row in examples:
        label = int(row["label"])
        if label not in by_label:
            raise ValueError(f"Unexpected label: {label}")
        by_label[label].append({"text": str(row["text"]), "label": label})
    rng = random.Random(seed)
    splits = {"train": [], "validation": [], "test": []}
    for label in (0, 1):
        rng.shuffle(by_label[label])
        allocated = _allocate_class(by_label[label])
        for name, rows in zip(splits, allocated, strict=True):
            splits[name].extend(rows)
    for rows in splits.values():
        rng.shuffle(rows)
    return splits


def compute_class_weights(examples: Sequence[Mapping[str, Any]]) -> list[float]:
    """Balanced weights n_samples / (n_classes * class_count)."""
    counts = Counter(int(row["label"]) for row in examples)
    if not counts.get(0) or not counts.get(1):
        raise InsufficientDataError("Class weights require both labels")
    total = counts[0] + counts[1]
    return [total / (2 * counts[0]), total / (2 * counts[1])]


def split_counts(splits: Mapping[str, Sequence[Mapping[str, Any]]]) -> dict[str, dict[str, int]]:
    return {
        name: {
            "NORMAL": sum(int(row["label"]) == 0 for row in rows),
            "SUSPICIOUS": sum(int(row["label"]) == 1 for row in rows),
        }
        for name, rows in splits.items()
    }


def read_review_master(path: str) -> list[dict[str, Any]]:
    """Read the reviewed workbook. openpyxl is kept out of runtime deps."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install the ml dependency group to read .xlsx files") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "REVIEW_MASTER" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a REVIEW_MASTER sheet")
    sheet = workbook["REVIEW_MASTER"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [dict(zip(headers, values, strict=True)) for values in rows]

